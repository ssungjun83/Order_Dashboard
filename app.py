from __future__ import annotations

import io
import re
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, Tuple

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_FILE = BASE_DIR / "order_status_with_leadtime.xlsx"
ISSUE_TRACKER_PATH = BASE_DIR / "issue_tracker.xlsx"

TAB_ORDER_STATUS = "\uc218\uc8fc \uc9c4\ud589 \uc0c1\uc138"
TAB_BY_ITEM = "\uc81c\ud488\ubcc4 \uc218\uc8fc \uc9c4\ud589"
TAB_ISSUES = "\uc0dd\uc0b0 \uc774\uc288 \uad00\ub9ac"
TAB_PRODUCT_SUMMARY = "\uc81c\ud488 \uc218\uc694 \uc694\uc57d"
SEARCH_COL = "__search_key__"
MAX_STATUS_STYLE_ROWS = 2000
ISSUE_ROW_HEIGHT = 90
ISSUE_TABLE_MAX_HEIGHT = 360
ISSUE_RESOLVED_MAX_HEIGHT = 260

THEME_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@300;400;600;700&family=IBM+Plex+Serif:wght@600;700&display=swap');
:root {
  --bg: #f6f3ea;
  --bg-2: #f2f5f7;
  --panel: #ffffff;
  --text: #1f2937;
  --muted: #6b7280;
  --accent: #e76f51;
  --accent-2: #2a9d8f;
  --border: #e7ddd0;
}
html, body, [class*="stApp"] {
  font-family: 'IBM Plex Sans', 'Noto Sans KR', 'Apple SD Gothic Neo', sans-serif;
  color: var(--text);
}
.stApp {
  background: radial-gradient(1200px 600px at 10% -10%, #f4e6d4 0%, var(--bg) 40%, var(--bg-2) 100%);
}
.block-container {
  padding-top: 2rem;
  animation: fadeInUp 0.4s ease-out;
}
@keyframes fadeInUp {
  from { opacity: 0; transform: translateY(8px); }
  to { opacity: 1; transform: translateY(0); }
}
h1, h2, h3 {
  font-family: 'IBM Plex Serif', 'Noto Serif KR', serif;
  letter-spacing: -0.5px;
}
h1 { font-size: 2.6rem; margin-bottom: 0.2rem; }
.stCaption { color: var(--muted); }
.stCaption { margin-bottom: 0.4rem; }
.stTabs [data-baseweb="tab"] {
  background: #fff8ee;
  border: 1px solid var(--border);
  border-radius: 999px;
  padding: 0.4rem 1rem;
  margin-right: 0.5rem;
}
.stTabs [data-baseweb="tab"][aria-selected="true"] {
  background: var(--accent);
  color: #ffffff;
  border-color: var(--accent);
}
.stTabs [data-baseweb="tab-list"] { gap: 0.35rem; border-bottom: none; }
.stSidebar > div {
  background: #fbf7f0;
  border-right: 1px solid var(--border);
}
.stTextInput input,
.stMultiSelect div[data-baseweb="select"] {
  background: #fffdf8;
  border-radius: 12px;
  border: 1px solid var(--border);
}
.stButton button,
.stDownloadButton button {
  background: #eeeeee;
  color: #1f2937;
  border-radius: 10px;
  padding: 0.25rem 0.65rem;
  border: 1px solid #d7d7d7;
  box-shadow: none;
  font-size: 0.78rem;
  line-height: 1.1;
  white-space: nowrap;
}
.stButton button:hover,
.stDownloadButton button:hover {
  background: #e7e7e7;
  border-color: #cfcfcf;
  color: #111827;
}
div[data-testid="stDataFrame"] {
  border: 1px solid var(--border);
  border-radius: 16px;
  padding: 0.5rem;
  background: #ffffff;
  box-shadow: 0 4px 16px rgba(16, 24, 40, 0.05);
  margin-top: 0.6rem;
  margin-bottom: 1.2rem;
}
div[data-testid="stDataFrame"] div[role="columnheader"] {
  background: #f8f2e7;
  font-weight: 600;
}
div[data-testid="stDataFrame"] div[role="row"] { background: #ffffff; }
div[data-testid="stDataFrame"] div[role="row"]:nth-child(even) { background: #fcfbf8; }
div[data-testid="stDownloadButton"] { margin-top: 0.35rem; }
div[data-testid="stMarkdownContainer"] { margin-bottom: 0.2rem; }
div[data-testid="stElementContainer"] {
  margin-bottom: 0.7rem;
}
div[data-testid="stElementContainer"] + div[data-testid="stElementContainer"] {
  margin-top: 0.25rem;
}
</style>
"""

STATUS_COLORS = {
    "\ucd9c\uace0\uc644\ub8cc": "#1e5db0",
    "\ud3ec\uc7a5\uc644\ub8cc": "#2f79c8",
    "\ud3ec\uc7a5\uc9c4\ud589\uc911": "#5aa0dc",
    "\uc0dd\uc0b0\uc644\ub8cc": "#8dbce6",
    "\uc0dd\uc0b0\uc9c4\ud589\uc911": "#ffffff",
    "\ubbf8\uc9c4\ud589": "#dcebfa",
}

COL_MONTH = "\uc6d4"
COL_TYPE = "\uad6c\ubd84"
COL_STATUS = "\ud604\uc7ac\uc0c1\ud0dc"
COL_COUNTRY = "\uad6d\uac00"
COL_OWNER = "\ub2f4\ub2f9\uc790"
COL_CUSTOMER = "\uace0\uac1d"
COL_WORKNO = "\uc791\uc9c0\ubc88\ud638"
COL_PRODUCT = "\ud488\uba85"
COL_YEAR = "\uc5f0\ub3c4"
COL_MONTH_DATE = "__month_date__"
COL_PROD_EXPECT = "\uc0dd\uc0b0\uc644\ub8cc\uc608\uc0c1\uc77c"

COL_ORDER_QTY = "\uc624\ub354\uc218\ub7c9"
COL_ORDER_AMT = "\uc218\uc8fc\uae08\uc561"
COL_ORDER_AMT_KRW = "\uc218\uc8fc\uae08\uc561(\uc6d0)"
COL_ORDER_AMT_USD = "\uc218\uc8fc\uae08\uc561(\ub2ec\ub7ec)"
COL_LEADTIME = "\ub9ac\ub4dc\ud0c0\uc784(\uc77c)"

COL_ORDER_SENT = "\uc218\uc8fc \uc804\uc1a1\uc77c"
COL_SALES_REQ = "\uc601\uc5c5\ucd9c\uace0\uc694\uccad\uc77c"
COL_FIRST_SHIP_PLAN = "\ucd5c\ucd08\ucd9c\uace0\uacc4\ud68d\uc77c"
COL_PACK_EXPECT = "\ud3ec\uc7a5\uc644\ub8cc\uc608\uc0c1\uc77c"
COL_PACK_DONE = "\ud3ec\uc7a5\uc644\ub8cc\uc77c"
COL_PACK_PROGRESS = "\ud3ec\uc7a5 \uc9c4\ub3c4\uc728"
COL_NOTE = "\uc0dd\uc0b0/\ud3ec\uc7a5 \ud2b9\uc774\uc0ac\ud56d"
COL_RESOLVED = "\ud574\uacb0\uc5ec\ubd80"
COL_CLOSED_DATE = "\uc885\uacb0\uc77c"
COL_ISSUE_DATE = "\uc548\uac74\uc0c1\uc815\uc77c"
COL_PRIORITY = "\uc6b0\uc120\uc21c\uc704"
COL_AVG_DEMAND = "\ud3c9\uade0\uc218\uc694"
COL_TOTAL_QTY = "\ucd1d \uc624\ub354\uc218\ub7c9"
COL_PO_COUNT = "PO\ud69f\uc218"
COL_PO_STREAK = "\uc5f0\uc18d PO \ud69f\uc218"
COL_SHARE = "\uc810\uc720\uc728"
COL_DUE_PLAN = "\ub0a9\uae30\uc900\uc218(\ucd5c\ucd08\ucd9c\uace0\uacc4\ud68d\uc77c)"
COL_DUE_SALES = "\ub0a9\uae30\uc900\uc218(\uc601\uc5c5\ucd9c\uace0\uc694\uccad\uc77c)"
COL_DUE_PLAN_RATE = "\ub0a9\uae30\uc900\uc218\uc728(\ucd5c\ucd08\ucd9c\uace0\uacc4\ud68d\uc77c)"
COL_ISSUE_KEY = "__issue_key__"

ORDER_STATUS_NUMERIC = [
    COL_ORDER_QTY,
    COL_ORDER_AMT,
    COL_ORDER_AMT_KRW,
    COL_ORDER_AMT_USD,
    COL_PACK_PROGRESS,
    COL_LEADTIME,
]
ORDER_STATUS_DATE = [
    COL_ORDER_SENT,
    COL_SALES_REQ,
    COL_FIRST_SHIP_PLAN,
    COL_PACK_EXPECT,
    COL_PACK_DONE,
]
ORDER_STATUS_MIXED_DATE = [COL_PROD_EXPECT]
ORDER_STATUS_PERCENT = [COL_PACK_PROGRESS]

MONTHLY_NUMERIC = [
    "\uc218\uc8fc\uac74\uc218",
    COL_ORDER_QTY,
    COL_ORDER_AMT,
    COL_ORDER_AMT_KRW,
    "\uc218\uc8fc\uae08\uc561(USD)",
]

LEADTIME_NUMERIC = [
    "\uc791\uc9c0\uac74\uc218",
    "\ub9ac\ub4dc\ud0c0\uc784\uac74\uc218",
    "\ud3c9\uade0\ub9ac\ub4dc\ud0c0\uc784(\uc77c)",
    "\ucd5c\uc18c\ub9ac\ub4dc\ud0c0\uc784(\uc77c)",
    "\ucd5c\ub300\ub9ac\ub4dc\ud0c0\uc784(\uc77c)",
]


def to_numeric(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    for col in columns:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    return df


def to_datetime(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    for col in columns:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")
    return df


def coerce_mixed_date(value: object) -> object:
    if pd.isna(value):
        return ""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, pd.Timestamp):
        return value.date()
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return ""
        parsed = pd.to_datetime(text, errors="coerce")
        if pd.notna(parsed):
            return parsed.date()
        return text
    return value


def prepare_display(
    df: pd.DataFrame,
    numeric_cols: list[str],
    date_cols: list[str],
    mixed_date_cols: list[str] | None = None,
    percent_cols: list[str] | None = None,
) -> pd.DataFrame:
    display = df.copy()
    percent_cols = percent_cols or []
    for col in numeric_cols:
        if col in display.columns:
            series = pd.to_numeric(display[col], errors="coerce")
            if col in percent_cols:
                display[col] = series.round(1)
            else:
                display[col] = series.round(0).astype("Int64")
    mixed_date_cols = mixed_date_cols or []
    for col in mixed_date_cols:
        if col in display.columns:
            display[col] = display[col].apply(coerce_mixed_date)
    for col in date_cols:
        if col in display.columns:
            display[col] = pd.to_datetime(display[col], errors="coerce").dt.date
    return display


def build_column_config(
    numeric_cols: list[str],
    date_cols: list[str],
    mixed_date_cols: list[str] | None = None,
    percent_cols: list[str] | None = None,
) -> Dict[str, st.column_config.ColumnConfig]:
    config: Dict[str, st.column_config.ColumnConfig] = {}
    percent_cols = percent_cols or []
    for col in numeric_cols:
        if col in percent_cols:
            config[col] = st.column_config.NumberColumn(format="%.1f%%")
        else:
            config[col] = st.column_config.NumberColumn(format="%,.0f")
    for col in date_cols:
        config[col] = st.column_config.DateColumn(format="iso8601")
    if mixed_date_cols:
        for col in mixed_date_cols:
            config[col] = st.column_config.TextColumn()
    return config


def column_width_hint(label: str) -> str:
    length = len(str(label))
    if length <= 4:
        return "small"
    if length <= 10:
        return "medium"
    return "large"


def build_width_config(
    columns: list[str],
    numeric_cols: list[str],
    percent_cols: list[str] | None = None,
) -> Dict[str, st.column_config.ColumnConfig]:
    config: Dict[str, st.column_config.ColumnConfig] = {}
    percent_cols = percent_cols or []
    numeric_set = set(numeric_cols)
    percent_set = set(percent_cols)
    for col in columns:
        width = column_width_hint(col)
        if col in numeric_set:
            if col in percent_set:
                config[col] = st.column_config.NumberColumn(
                    format="%.1f%%", width=width
                )
            else:
                config[col] = st.column_config.NumberColumn(
                    format="%,.0f", width=width
                )
        else:
            config[col] = st.column_config.TextColumn(width=width)
    return config


def apply_styler_widths(
    styled: pd.io.formats.style.Styler, columns: list[str]
) -> pd.io.formats.style.Styler:
    styles = []
    for idx, col in enumerate(columns):
        length = len(str(col))
        if length <= 4:
            width = "90px"
        elif length <= 10:
            width = "140px"
        else:
            width = "200px"
        styles.append(
            {
                "selector": f"th.col{idx}",
                "props": [("min-width", width), ("max-width", width)],
            }
        )
        styles.append(
            {
                "selector": f"td.col{idx}",
                "props": [("min-width", width), ("max-width", width)],
            }
        )
    return styled.set_table_styles(styles, overwrite=False)


def format_number(value: object) -> str:
    if pd.isna(value):
        return ""
    try:
        return f"{int(round(float(value))):,}"
    except (TypeError, ValueError):
        return str(value)


def format_percent(value: object) -> str:
    if pd.isna(value):
        return ""
    try:
        return f"{float(value):.1f}%"
    except (TypeError, ValueError):
        return str(value)


def format_date(value: object) -> str:
    if pd.isna(value):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, pd.Timestamp):
        return value.date().isoformat()
    text = str(value)
    return text[:10] if len(text) >= 10 else text


def style_status(value: object) -> str:
    text = "" if pd.isna(value) else str(value)
    color = STATUS_COLORS.get(text, "#e6f1fb")
    text_color = "#ffffff" if text == "\ucd9c\uace0\uc644\ub8cc" else "#0f172a"
    return f"background-color: {color}; color: {text_color}; font-weight: 600;"


def style_total_row(row: pd.Series) -> list[str]:
    if COL_TYPE in row.index and str(row[COL_TYPE]) == "\ud569\uacc4":
        return ["font-weight: 700;"] * len(row)
    return [""] * len(row)


def build_styler(
    df: pd.DataFrame,
    numeric_cols: list[str],
    date_cols: list[str],
    status_col: str | None = None,
    percent_cols: list[str] | None = None,
) -> pd.io.formats.style.Styler:
    percent_cols = percent_cols or []
    percent = [col for col in percent_cols if col in df.columns]
    numeric = [
        col for col in numeric_cols if col in df.columns and col not in percent
    ]
    date_list = [col for col in date_cols if col in df.columns]
    text = [col for col in df.columns if col not in set(numeric + date_list + percent)]
    if status_col and status_col in text:
        text = [col for col in text if col != status_col]

    styled = df.style
    right_align = numeric + percent
    if right_align:
        styled = styled.set_properties(subset=right_align, **{"text-align": "right"})
    if date_list:
        styled = styled.set_properties(subset=date_list, **{"text-align": "center"})
    if text:
        styled = styled.set_properties(subset=text, **{"text-align": "left"})
    if status_col and status_col in df.columns:
        styled = styled.set_properties(
            subset=[status_col], **{"text-align": "center"}
        )
    styled = styled.set_table_styles(
        [{"selector": "th", "props": [("text-align", "center")]}]
    )

    formatters: Dict[str, object] = {}
    for col in numeric:
        formatters[col] = format_number
    for col in percent:
        formatters[col] = format_percent
    if formatters:
        styled = styled.format(formatters, na_rep="")

    if status_col and status_col in df.columns:
        styled = styled.applymap(style_status, subset=[status_col])

    return styled


def add_year_column(df: pd.DataFrame) -> pd.DataFrame:
    if COL_MONTH not in df.columns:
        return df
    years = (
        df[COL_MONTH]
        .astype(str)
        .str.extract(r"(?P<yy>\d{2})\.", expand=True)["yy"]
        .astype("float")
        .apply(lambda v: int(2000 + v) if pd.notna(v) else None)
    )
    df[COL_YEAR] = years
    return df


def add_month_date_column(df: pd.DataFrame) -> pd.DataFrame:
    if COL_MONTH_DATE in df.columns or COL_MONTH not in df.columns:
        return df
    extracted = df[COL_MONTH].astype(str).str.extract(r"(?P<yy>\d{2})\.(?P<mm>\d{2})")
    years = pd.to_numeric(extracted["yy"], errors="coerce")
    months = pd.to_numeric(extracted["mm"], errors="coerce")
    month_date = pd.to_datetime(
        {"year": years + 2000, "month": months, "day": 1}, errors="coerce"
    ).dt.date
    df = df.copy()
    df[COL_MONTH_DATE] = month_date
    return df


def compute_compliance_rate(
    df: pd.DataFrame, group_cols: list[str]
) -> pd.Series:
    if COL_DUE_PLAN not in df.columns:
        return pd.Series(dtype="float")
    total = df.groupby(group_cols, dropna=False)[COL_WORKNO].count()
    delayed = (
        df[df[COL_DUE_PLAN] == "\uc9c0\uc5f0"]
        .groupby(group_cols, dropna=False)[COL_WORKNO]
        .count()
    )
    delayed = delayed.reindex(total.index, fill_value=0)
    rate = (total - delayed).div(total).mul(100)
    return rate


def max_consecutive_months(months: list[date]) -> int:
    if not months:
        return 0
    uniq = sorted({m for m in months if isinstance(m, date)})
    if not uniq:
        return 0
    best = 1
    current = 1
    prev_key = uniq[0].year * 12 + uniq[0].month
    for current_date in uniq[1:]:
        cur_key = current_date.year * 12 + current_date.month
        if cur_key - prev_key == 1:
            current += 1
        else:
            current = 1
        best = max(best, current)
        prev_key = cur_key
    return best


def compute_product_priority(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or COL_PRODUCT not in df.columns or COL_ORDER_QTY not in df.columns:
        return pd.DataFrame(
            columns=[COL_PRIORITY, COL_PRODUCT, COL_AVG_DEMAND, COL_PO_COUNT, COL_PO_STREAK, COL_SHARE]
        )
    df = add_month_date_column(df.copy())
    df = df[df[COL_PRODUCT].notna()]
    df = df[df[COL_PRODUCT].astype(str).str.strip().ne("")]

    total_qty = df[COL_ORDER_QTY].sum()
    grouped = df.groupby(COL_PRODUCT, dropna=False)
    summary = grouped.agg(
        _total_qty=(COL_ORDER_QTY, "sum"),
        _po_count=(COL_WORKNO, "nunique"),
        _avg_demand=(COL_ORDER_QTY, "mean"),
    )
    summary[COL_PO_STREAK] = grouped[COL_MONTH_DATE].apply(
        lambda values: max_consecutive_months([v for v in values if pd.notna(v)])
    )
    if total_qty:
        summary[COL_SHARE] = (summary["_total_qty"] / total_qty) * 100
    else:
        summary[COL_SHARE] = 0.0

    summary = summary.sort_values(
        by=[COL_SHARE, "_avg_demand", COL_PO_STREAK],
        ascending=[False, False, False],
    )
    summary[COL_PRIORITY] = range(1, len(summary) + 1)
    summary[COL_AVG_DEMAND] = summary["_avg_demand"]
    summary[COL_TOTAL_QTY] = summary["_total_qty"]
    summary[COL_PO_COUNT] = summary["_po_count"]
    summary = summary.reset_index()
    summary = summary[
        [
            COL_PRIORITY,
            COL_PRODUCT,
            COL_AVG_DEMAND,
            COL_TOTAL_QTY,
            COL_PO_COUNT,
            COL_PO_STREAK,
            COL_SHARE,
        ]
    ]
    return summary.reset_index(drop=True)


def move_note_before_year(df: pd.DataFrame) -> pd.DataFrame:
    if COL_NOTE not in df.columns or COL_YEAR not in df.columns:
        return df
    cols = list(df.columns)
    cols.remove(COL_NOTE)
    year_idx = cols.index(COL_YEAR)
    cols.insert(year_idx, COL_NOTE)
    return df[cols]


def replace_capa_delay(df: pd.DataFrame) -> pd.DataFrame:
    return df


def build_issue_key(df: pd.DataFrame) -> pd.Series:
    return (
        df[[COL_WORKNO, COL_CUSTOMER, COL_PRODUCT, COL_NOTE]]
        .fillna("")
        .astype(str)
        .agg("|".join, axis=1)
    )


def load_issue_tracker(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame(
            columns=[COL_ISSUE_KEY, COL_RESOLVED, COL_CLOSED_DATE, COL_ISSUE_DATE]
        )
    df = pd.read_excel(path)
    if COL_ISSUE_KEY not in df.columns:
        return pd.DataFrame(
            columns=[COL_ISSUE_KEY, COL_RESOLVED, COL_CLOSED_DATE, COL_ISSUE_DATE]
        )
    if COL_RESOLVED not in df.columns:
        df[COL_RESOLVED] = False
    if COL_CLOSED_DATE not in df.columns:
        df[COL_CLOSED_DATE] = pd.NaT
    if COL_ISSUE_DATE not in df.columns:
        df[COL_ISSUE_DATE] = pd.NaT
    df[COL_RESOLVED] = df[COL_RESOLVED].fillna(False).astype(bool)
    df[COL_CLOSED_DATE] = pd.to_datetime(df[COL_CLOSED_DATE], errors="coerce").dt.date
    df[COL_ISSUE_DATE] = pd.to_datetime(df[COL_ISSUE_DATE], errors="coerce").dt.date
    return df[[COL_ISSUE_KEY, COL_RESOLVED, COL_CLOSED_DATE, COL_ISSUE_DATE]]


def save_issue_tracker(df: pd.DataFrame, path: Path) -> None:
    export_df = df[
        [COL_ISSUE_KEY, COL_RESOLVED, COL_CLOSED_DATE, COL_ISSUE_DATE]
    ].copy()
    export_df.to_excel(path, index=False)


def add_months(base: date, offset: int) -> date:
    total = base.month - 1 + offset
    year = base.year + total // 12
    month = total % 12 + 1
    return date(year, month, 1)


def last_day_of_month(value: date) -> date:
    next_month = add_months(value.replace(day=1), 1)
    return next_month - timedelta(days=1)


def start_of_quarter(value: date) -> date:
    quarter = (value.month - 1) // 3
    return date(value.year, quarter * 3 + 1, 1)


def end_of_quarter(value: date) -> date:
    return last_day_of_month(add_months(start_of_quarter(value), 2))


def start_of_half_year(value: date) -> date:
    return date(value.year, 1, 1) if value.month <= 6 else date(value.year, 7, 1)


def end_of_half_year(value: date) -> date:
    return date(value.year, 6, 30) if value.month <= 6 else date(value.year, 12, 31)


def clamp_range(start: date, end: date, min_date: date, max_date: date) -> Tuple[date, date]:
    if start < min_date:
        start = min_date
    if end > max_date:
        end = max_date
    if end < start:
        end = start
    return start, end


def render_period_controls(df: pd.DataFrame, key_prefix: str) -> Tuple[date, date] | None:
    month_range = None
    if COL_MONTH_DATE not in df.columns:
        return month_range
    month_values = df[COL_MONTH_DATE].dropna().tolist()
    if not month_values:
        return month_range

    month_min = min(month_values)
    month_max = max(month_values)
    min_date = month_min
    max_date = last_day_of_month(month_max)
    current_month = date.today().replace(day=1)
    default_start = current_month
    default_end = last_day_of_month(add_months(current_month, 2))
    default_start, default_end = clamp_range(default_start, default_end, min_date, max_date)

    period_key = f"{key_prefix}_period_range"
    if period_key not in st.session_state:
        st.session_state[period_key] = (default_start, default_end)

    today = date.today()
    presets = {
        "\ub2f9\uc6d4": (date(today.year, today.month, 1), last_day_of_month(today)),
        "\ub2f9\ubd84\uae30": (start_of_quarter(today), end_of_quarter(today)),
        "\ub2f9\ubc18\uae30": (start_of_half_year(today), end_of_half_year(today)),
        "\uc62c\ud574": (date(today.year, 1, 1), date(today.year, 12, 31)),
        "\uc791\ub144": (date(today.year - 1, 1, 1), date(today.year - 1, 12, 31)),
        "2\uac1c\ub144": (date(today.year - 1, 1, 1), date(today.year, 12, 31)),
        "3\uac1c\ub144": (date(today.year - 2, 1, 1), date(today.year, 12, 31)),
        "\ucd5c\ub300\uae30\uac04": (min_date, max_date),
    }

    col_period, col_presets = st.columns([4, 8])
    with col_presets:
        btn_cols = st.columns([4, 1, 1, 1, 1, 1, 1, 1, 1], gap="small")
        labels = [
            "\ub2f9\uc6d4",
            "\ub2f9\ubd84\uae30",
            "\ub2f9\ubc18\uae30",
            "\uc62c\ud574",
            "\uc791\ub144",
            "2\uac1c\ub144",
            "3\uac1c\ub144",
            "\ucd5c\ub300\uae30\uac04",
        ]
        for idx, label in enumerate(labels, start=1):
            start, end = presets[label]
            start, end = clamp_range(start, end, min_date, max_date)
            with btn_cols[idx]:
                if st.button(label, key=f"{key_prefix}_preset_{idx}", use_container_width=True):
                    st.session_state[period_key] = (start, end)

    with col_period:
        selected = st.date_input(
            "\uae30\uac04",
            key=period_key,
            min_value=min_date,
            max_value=max_date,
        )

    if isinstance(selected, tuple):
        start, end = selected
    else:
        start = selected
        end = selected
    start, end = clamp_range(start, end, min_date, max_date)
    month_range = (date(start.year, start.month, 1), date(end.year, end.month, 1))
    return month_range


def calc_table_height(
    row_count: int, row_height: int = 34, header_height: int = 36, max_height: int = 360
) -> int:
    if row_count <= 0:
        return header_height + row_height
    return min(max_height, header_height + row_height * row_count)


def add_search_column(df: pd.DataFrame) -> pd.DataFrame:
    if SEARCH_COL in df.columns:
        return df
    text = df.fillna("").astype(str).agg(" ".join, axis=1).str.lower()
    df = df.copy()
    df[SEARCH_COL] = text
    return df


@st.cache_data(show_spinner=False)
def load_from_path(path: str, mtime: float) -> Dict[str, pd.DataFrame]:
    xl = pd.ExcelFile(path)
    data = {
        "order_status": pd.read_excel(xl, sheet_name="order_status"),
        "order_status_by_item": pd.read_excel(xl, sheet_name="order_status_by_item"),
        "monthly_summary": pd.read_excel(xl, sheet_name="monthly_summary"),
        "summary_by_month": pd.read_excel(xl, sheet_name="summary_by_month"),
    }
    data["order_status"] = to_datetime(
        to_numeric(data["order_status"], ORDER_STATUS_NUMERIC), ORDER_STATUS_DATE
    )
    data["order_status"] = replace_capa_delay(data["order_status"])
    data["order_status"] = add_year_column(data["order_status"])
    data["order_status"] = add_month_date_column(data["order_status"])
    data["order_status"] = add_search_column(data["order_status"])
    data["order_status_by_item"] = to_datetime(
        to_numeric(data["order_status_by_item"], ORDER_STATUS_NUMERIC), ORDER_STATUS_DATE
    )
    data["order_status_by_item"] = replace_capa_delay(data["order_status_by_item"])
    data["order_status_by_item"] = add_year_column(data["order_status_by_item"])
    data["order_status_by_item"] = add_month_date_column(data["order_status_by_item"])
    data["order_status_by_item"] = add_search_column(data["order_status_by_item"])
    data["monthly_summary"] = to_numeric(data["monthly_summary"], MONTHLY_NUMERIC)
    data["summary_by_month"] = to_numeric(data["summary_by_month"], LEADTIME_NUMERIC)
    return data


@st.cache_data(show_spinner=False)
def load_from_bytes(content: bytes, key: str) -> Dict[str, pd.DataFrame]:
    xl = pd.ExcelFile(io.BytesIO(content))
    data = {
        "order_status": pd.read_excel(xl, sheet_name="order_status"),
        "order_status_by_item": pd.read_excel(xl, sheet_name="order_status_by_item"),
        "monthly_summary": pd.read_excel(xl, sheet_name="monthly_summary"),
        "summary_by_month": pd.read_excel(xl, sheet_name="summary_by_month"),
    }
    data["order_status"] = to_datetime(
        to_numeric(data["order_status"], ORDER_STATUS_NUMERIC), ORDER_STATUS_DATE
    )
    data["order_status"] = replace_capa_delay(data["order_status"])
    data["order_status"] = add_year_column(data["order_status"])
    data["order_status"] = add_month_date_column(data["order_status"])
    data["order_status"] = add_search_column(data["order_status"])
    data["order_status_by_item"] = to_datetime(
        to_numeric(data["order_status_by_item"], ORDER_STATUS_NUMERIC), ORDER_STATUS_DATE
    )
    data["order_status_by_item"] = replace_capa_delay(data["order_status_by_item"])
    data["order_status_by_item"] = add_year_column(data["order_status_by_item"])
    data["order_status_by_item"] = add_month_date_column(data["order_status_by_item"])
    data["order_status_by_item"] = add_search_column(data["order_status_by_item"])
    data["monthly_summary"] = to_numeric(data["monthly_summary"], MONTHLY_NUMERIC)
    data["summary_by_month"] = to_numeric(data["summary_by_month"], LEADTIME_NUMERIC)
    return data


def apply_order_filters(
    df: pd.DataFrame,
    month_range: Tuple[date, date] | None,
    filters: dict | None = None,
    show_sidebar: bool = True,
    apply_month_filter: bool = True,
) -> Tuple[pd.DataFrame, pd.DataFrame, dict]:
    if filters is None:
        filters = {}

    if show_sidebar:
        with st.sidebar:
            st.subheader("\ud544\ud130")
            df_period = df
            if month_range and COL_MONTH_DATE in df.columns:
                start, end = month_range
                df_period = df[
                    (df[COL_MONTH_DATE] >= start) & (df[COL_MONTH_DATE] <= end)
                ]

            months = sorted(df_period[COL_MONTH].dropna().unique().tolist())
            types = sorted(df_period[COL_TYPE].dropna().unique().tolist())
            statuses = sorted(
                df_period.get(COL_STATUS, pd.Series(dtype=str)).dropna().unique().tolist()
            )
            countries = sorted(
                df_period.get(COL_COUNTRY, pd.Series(dtype=str)).dropna().unique().tolist()
            )
            owners = sorted(
                df_period.get(COL_OWNER, pd.Series(dtype=str)).dropna().unique().tolist()
            )
            customers = sorted(
                df_period.get(COL_CUSTOMER, pd.Series(dtype=str)).dropna().unique().tolist()
            )

            filters["months"] = st.multiselect(COL_MONTH, months, default=months)
            filters["types"] = st.multiselect(COL_TYPE, types, default=types)
            filters["statuses"] = st.multiselect(COL_STATUS, statuses, default=statuses)
            filters["countries"] = st.multiselect(COL_COUNTRY, countries, default=countries)
            filters["owners"] = st.multiselect(COL_OWNER, owners, default=owners)
            filters["customers"] = st.multiselect(COL_CUSTOMER, customers, default=customers)

    filters.setdefault("months", [])
    filters.setdefault("types", [])
    filters.setdefault("statuses", [])
    filters.setdefault("countries", [])
    filters.setdefault("owners", [])
    filters.setdefault("customers", [])

    base_df = df
    if apply_month_filter and filters["months"]:
        base_df = base_df[base_df[COL_MONTH].isin(filters["months"])]
    if filters["types"]:
        base_df = base_df[base_df[COL_TYPE].isin(filters["types"])]
    if filters["statuses"] and COL_STATUS in base_df.columns:
        base_df = base_df[base_df[COL_STATUS].isin(filters["statuses"])]
    if filters["countries"] and COL_COUNTRY in base_df.columns:
        base_df = base_df[base_df[COL_COUNTRY].isin(filters["countries"])]
    if filters["owners"] and COL_OWNER in base_df.columns:
        base_df = base_df[base_df[COL_OWNER].isin(filters["owners"])]
    if filters["customers"] and COL_CUSTOMER in base_df.columns:
        base_df = base_df[base_df[COL_CUSTOMER].isin(filters["customers"])]

    if month_range and COL_MONTH_DATE in base_df.columns:
        start, end = month_range
        base_df = base_df[(base_df[COL_MONTH_DATE] >= start) & (base_df[COL_MONTH_DATE] <= end)]

    return base_df, base_df, filters


def apply_search(df: pd.DataFrame, query: str) -> pd.DataFrame:
    query = query.strip()
    if not query:
        return df
    if "," in query:
        suffix_patterns = [
            r",\s*co\.?\s*,\s*ltd\.?",
            r",\s*co\.?\s*ltd\.?",
            r",\s*ltd\.?",
            r",\s*inc\.?",
            r",\s*corp\.?",
            r",\s*llc\.?",
            r",\s*plc\.?",
            r",\s*gmbh\.?",
            r",\s*sa\.?",
            r",\s*srl\.?",
            r",\s*bv\.?",
            r",\s*kg\.?",
        ]
        temp = query.lower()
        for pattern in suffix_patterns:
            temp = re.sub(pattern, "", temp)
        if "," in temp:
            parts = [part.strip() for part in query.split(",") if part.strip()]
        else:
            parts = [query]
    else:
        parts = [query]
    groups = []
    for part in parts:
        tokens = [t.lower() for t in part.split() if t]
        if tokens:
            groups.append(tokens)
    if not groups:
        return df

    if SEARCH_COL in df.columns:
        text_series = df[SEARCH_COL].astype(str)
        mask = pd.Series(False, index=df.index)
        for tokens in groups:
            group_mask = pd.Series(True, index=df.index)
            for token in tokens:
                group_mask &= text_series.str.contains(token, na=False)
            mask |= group_mask
        return df[mask]

    text_df = df.astype(str).apply(lambda s: s.str.lower())
    mask = pd.Series(False, index=df.index)
    for tokens in groups:
        group_mask = pd.Series(True, index=df.index)
        for token in tokens:
            term_mask = pd.Series(False, index=df.index)
            for col in text_df.columns:
                term_mask |= text_df[col].str.contains(token, na=False)
            group_mask &= term_mask
        mask |= group_mask
    return df[mask]


def render_year_summary(df: pd.DataFrame, key_prefix: str) -> None:
    if COL_YEAR not in df.columns:
        return
    show_monthly = st.toggle(
        "\uc6d4\ubcc4 \uc694\uc57d \ubcf4\uae30",
        value=False,
        key=f"{key_prefix}_monthly_toggle",
    )
    summary_type = (
        df.groupby([COL_YEAR, COL_TYPE], dropna=False)
        .agg(
            **{
                "\uc791\uc9c0\uac74\uc218": (COL_WORKNO, "count"),
                "\uc624\ub354\uc218\ub7c9\ud569\uacc4": (COL_ORDER_QTY, "sum"),
                "\uc218\uc8fc\uae08\uc561\ud569\uacc4": (COL_ORDER_AMT, "sum"),
                "\uc218\uc8fc\uae08\uc561\uc6d0\ud569\uacc4": (COL_ORDER_AMT_KRW, "sum"),
                "\uc218\uc8fc\uae08\uc561\ub2ec\ub7ec\ud569\uacc4": (COL_ORDER_AMT_USD, "sum"),
                "\ud3c9\uade0\ub9ac\ub4dc\ud0c0\uc784\uc77c": (COL_LEADTIME, "mean"),
            }
        )
        .reset_index()
    )
    rate_type = compute_compliance_rate(df, [COL_YEAR, COL_TYPE]).reset_index(
        name=COL_DUE_PLAN_RATE
    )
    summary_type = summary_type.merge(rate_type, on=[COL_YEAR, COL_TYPE], how="left")
    summary_total = (
        df.groupby(COL_YEAR, dropna=False)
        .agg(
            **{
                "\uc791\uc9c0\uac74\uc218": (COL_WORKNO, "count"),
                "\uc624\ub354\uc218\ub7c9\ud569\uacc4": (COL_ORDER_QTY, "sum"),
                "\uc218\uc8fc\uae08\uc561\ud569\uacc4": (COL_ORDER_AMT, "sum"),
                "\uc218\uc8fc\uae08\uc561\uc6d0\ud569\uacc4": (COL_ORDER_AMT_KRW, "sum"),
                "\uc218\uc8fc\uae08\uc561\ub2ec\ub7ec\ud569\uacc4": (COL_ORDER_AMT_USD, "sum"),
                "\ud3c9\uade0\ub9ac\ub4dc\ud0c0\uc784\uc77c": (COL_LEADTIME, "mean"),
            }
        )
        .reset_index()
    )
    rate_total = compute_compliance_rate(df, [COL_YEAR]).reset_index(
        name=COL_DUE_PLAN_RATE
    )
    summary_total = summary_total.merge(rate_total, on=[COL_YEAR], how="left")
    summary_total[COL_TYPE] = "\ud569\uacc4"
    summary = pd.concat([summary_type, summary_total], ignore_index=True)
    summary["__type_order"] = summary[COL_TYPE].apply(
        lambda value: 2 if value == "\ud569\uacc4" else 1
    )
    summary = summary.sort_values([COL_YEAR, "__type_order", COL_TYPE]).drop(
        columns="__type_order"
    )
    display = summary.rename(
        columns={
            COL_YEAR: COL_YEAR,
            COL_TYPE: COL_TYPE,
            "\uc791\uc9c0\uac74\uc218": "\uc791\uc9c0\uac74\uc218",
            "\uc624\ub354\uc218\ub7c9\ud569\uacc4": "\uc624\ub354\uc218\ub7c9 \ud569\uacc4",
            "\uc218\uc8fc\uae08\uc561\ud569\uacc4": "\uc218\uc8fc\uae08\uc561 \ud569\uacc4",
            "\uc218\uc8fc\uae08\uc561\uc6d0\ud569\uacc4": "\uc218\uc8fc\uae08\uc561(\uc6d0) \ud569\uacc4",
            "\uc218\uc8fc\uae08\uc561\ub2ec\ub7ec\ud569\uacc4": "\uc218\uc8fc\uae08\uc561(\ub2ec\ub7ec) \ud569\uacc4",
            "\ud3c9\uade0\ub9ac\ub4dc\ud0c0\uc784\uc77c": "\ud3c9\uade0 \ub9ac\ub4dc\ud0c0\uc784(\uc77c)",
        }
    )
    display[COL_YEAR] = display[COL_YEAR].fillna(0).astype("Int64")
    display[COL_YEAR] = display[COL_YEAR].astype(object)
    display.loc[
        display.duplicated(subset=[COL_YEAR], keep="first"), COL_YEAR
    ] = ""
    st.subheader("\ub144\ub3c4\ubcc4 \uc694\uc57d")
    numeric_cols = [
        COL_YEAR,
        "\uc791\uc9c0\uac74\uc218",
        "\uc624\ub354\uc218\ub7c9 \ud569\uacc4",
        "\uc218\uc8fc\uae08\uc561 \ud569\uacc4",
        "\uc218\uc8fc\uae08\uc561(\uc6d0) \ud569\uacc4",
        "\uc218\uc8fc\uae08\uc561(\ub2ec\ub7ec) \ud569\uacc4",
        "\ud3c9\uade0 \ub9ac\ub4dc\ud0c0\uc784(\uc77c)",
        COL_DUE_PLAN_RATE,
    ]
    display = prepare_display(display, numeric_cols, [], [], [COL_DUE_PLAN_RATE])
    styled = build_styler(display, numeric_cols, [], percent_cols=[COL_DUE_PLAN_RATE])
    styled = styled.apply(style_total_row, axis=1)
    styled = apply_styler_widths(styled, list(display.columns))
    st.dataframe(
        styled,
        use_container_width=True,
        height=calc_table_height(len(display), max_height=220),
    )

    if not show_monthly:
        return

    monthly_type = (
        df.groupby([COL_YEAR, COL_MONTH, COL_TYPE], dropna=False)
        .agg(
            **{
                "\uc791\uc9c0\uac74\uc218": (COL_WORKNO, "count"),
                "\uc624\ub354\uc218\ub7c9\ud569\uacc4": (COL_ORDER_QTY, "sum"),
                "\uc218\uc8fc\uae08\uc561\ud569\uacc4": (COL_ORDER_AMT, "sum"),
                "\uc218\uc8fc\uae08\uc561\uc6d0\ud569\uacc4": (COL_ORDER_AMT_KRW, "sum"),
                "\uc218\uc8fc\uae08\uc561\ub2ec\ub7ec\ud569\uacc4": (COL_ORDER_AMT_USD, "sum"),
                "\ud3c9\uade0\ub9ac\ub4dc\ud0c0\uc784\uc77c": (COL_LEADTIME, "mean"),
            }
        )
        .reset_index()
    )
    monthly_rate_type = compute_compliance_rate(
        df, [COL_YEAR, COL_MONTH, COL_TYPE]
    ).reset_index(name=COL_DUE_PLAN_RATE)
    monthly_type = monthly_type.merge(
        monthly_rate_type, on=[COL_YEAR, COL_MONTH, COL_TYPE], how="left"
    )
    monthly_total = (
        df.groupby([COL_YEAR, COL_MONTH], dropna=False)
        .agg(
            **{
                "\uc791\uc9c0\uac74\uc218": (COL_WORKNO, "count"),
                "\uc624\ub354\uc218\ub7c9\ud569\uacc4": (COL_ORDER_QTY, "sum"),
                "\uc218\uc8fc\uae08\uc561\ud569\uacc4": (COL_ORDER_AMT, "sum"),
                "\uc218\uc8fc\uae08\uc561\uc6d0\ud569\uacc4": (COL_ORDER_AMT_KRW, "sum"),
                "\uc218\uc8fc\uae08\uc561\ub2ec\ub7ec\ud569\uacc4": (COL_ORDER_AMT_USD, "sum"),
                "\ud3c9\uade0\ub9ac\ub4dc\ud0c0\uc784\uc77c": (COL_LEADTIME, "mean"),
            }
        )
        .reset_index()
    )
    monthly_rate_total = compute_compliance_rate(
        df, [COL_YEAR, COL_MONTH]
    ).reset_index(name=COL_DUE_PLAN_RATE)
    monthly_total = monthly_total.merge(
        monthly_rate_total, on=[COL_YEAR, COL_MONTH], how="left"
    )
    monthly_total[COL_TYPE] = "\ud569\uacc4"
    monthly = pd.concat([monthly_type, monthly_total], ignore_index=True)
    monthly["__type_order"] = monthly[COL_TYPE].apply(
        lambda value: 2 if value == "\ud569\uacc4" else 1
    )
    monthly = monthly.sort_values([COL_YEAR, COL_MONTH, "__type_order", COL_TYPE]).drop(
        columns="__type_order"
    )
    monthly = monthly.rename(
        columns={
            COL_YEAR: COL_YEAR,
            COL_MONTH: COL_MONTH,
            COL_TYPE: COL_TYPE,
            "\uc791\uc9c0\uac74\uc218": "\uc791\uc9c0\uac74\uc218",
            "\uc624\ub354\uc218\ub7c9\ud569\uacc4": "\uc624\ub354\uc218\ub7c9 \ud569\uacc4",
            "\uc218\uc8fc\uae08\uc561\ud569\uacc4": "\uc218\uc8fc\uae08\uc561 \ud569\uacc4",
            "\uc218\uc8fc\uae08\uc561\uc6d0\ud569\uacc4": "\uc218\uc8fc\uae08\uc561(\uc6d0) \ud569\uacc4",
            "\uc218\uc8fc\uae08\uc561\ub2ec\ub7ec\ud569\uacc4": "\uc218\uc8fc\uae08\uc561(\ub2ec\ub7ec) \ud569\uacc4",
            "\ud3c9\uade0\ub9ac\ub4dc\ud0c0\uc784\uc77c": "\ud3c9\uade0 \ub9ac\ub4dc\ud0c0\uc784(\uc77c)",
        }
    )
    monthly[COL_YEAR] = monthly[COL_YEAR].fillna(0).astype("Int64")
    monthly[COL_YEAR] = monthly[COL_YEAR].astype(object)
    monthly.loc[
        monthly.duplicated(subset=[COL_YEAR], keep="first"), COL_YEAR
    ] = ""
    monthly_display = prepare_display(monthly, numeric_cols, [], [], [COL_DUE_PLAN_RATE])
    monthly_styled = build_styler(
        monthly_display, numeric_cols, [], percent_cols=[COL_DUE_PLAN_RATE]
    )
    monthly_styled = monthly_styled.apply(style_total_row, axis=1)
    monthly_styled = apply_styler_widths(
        monthly_styled, list(monthly_display.columns)
    )
    st.subheader("\uc6d4\ubcc4 \uc694\uc57d")
    st.dataframe(
        monthly_styled,
        use_container_width=True,
        height=calc_table_height(len(monthly_display), max_height=360),
    )


def download_excel_button(
    df: pd.DataFrame,
    filename: str,
    numeric_cols: list[str],
    date_cols: list[str],
    key: str,
    mixed_date_cols: list[str] | None = None,
    percent_cols: list[str] | None = None,
) -> None:
    export_df = prepare_display(
        df, numeric_cols, date_cols, mixed_date_cols, percent_cols
    )
    if mixed_date_cols:
        for col in mixed_date_cols:
            if col not in export_df.columns:
                continue
            export_df[col] = export_df[col].apply(
                lambda v: v.isoformat()
                if isinstance(v, (date, datetime, pd.Timestamp))
                else ("" if v is None else str(v))
            )
    export_df = export_df.astype(object).where(pd.notna(export_df), None)
    buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    columns = list(export_df.columns)
    sum_exclude = {COL_YEAR, COL_LEADTIME, COL_PACK_PROGRESS}
    percent_set = set(percent_cols or [])
    numeric_set = set(numeric_cols) - percent_set
    date_set = set(date_cols + (mixed_date_cols or []))
    sum_cols = [
        col for col in numeric_set if col in columns and col not in sum_exclude
    ]
    data_start_row = 3
    data_end_row = data_start_row + len(export_df) - 1

    sum_row = []
    for col in columns:
        if col in sum_cols and len(export_df) > 0:
            col_letter = get_column_letter(columns.index(col) + 1)
            sum_row.append(
                f"=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})"
            )
        else:
            sum_row.append(None)
    ws.append(sum_row)
    ws.append(columns)
    for row in export_df.itertuples(index=False, name=None):
        ws.append(list(row))

    if len(export_df) > 0:
        for row in ws.iter_rows(min_row=data_start_row, max_row=data_end_row):
            for cell in row:
                col_name = columns[cell.column - 1]
                if col_name in date_set and isinstance(cell.value, datetime):
                    cell.value = cell.value.date()
                if col_name in percent_set and isinstance(cell.value, (int, float)):
                    cell.number_format = '0.0"%"'
                elif col_name in numeric_set and isinstance(cell.value, (int, float)):
                    cell.number_format = "#,###"
                elif col_name in date_set and isinstance(cell.value, date):
                    cell.number_format = "yyyy-mm-dd"

    for idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx)
        if col_name in numeric_set:
            cell.number_format = "#,###"
            cell.font = cell.font.copy(bold=True)
    for idx in range(1, len(columns) + 1):
        header_cell = ws.cell(row=2, column=idx)
        header_cell.font = header_cell.font.copy(bold=True)

    widths = {}
    sum_values = {}
    if sum_cols:
        sum_frame = export_df[sum_cols].apply(pd.to_numeric, errors="coerce")
        sum_values = sum_frame.sum().to_dict()
    for col_name in columns:
        max_len = len(str(col_name))
        if col_name in percent_set:
            formatter = format_percent
        elif col_name in numeric_set:
            formatter = format_number
        elif col_name in date_set:
            formatter = format_date
        else:
            formatter = lambda v: "" if pd.isna(v) else str(v)
        for value in export_df[col_name].tolist():
            text = formatter(value)
            max_len = max(max_len, len(text))
        if col_name in sum_values:
            text = formatter(sum_values[col_name])
            max_len = max(max_len, len(text))
        widths[col_name] = max_len + 2
    for idx, col_name in enumerate(columns, start=1):
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = widths.get(col_name, 10)

    ws.sheet_view.showGridLines = False
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for row in ws.iter_rows(min_row=2, max_row=data_end_row if len(export_df) else 2):
        for cell in row:
            cell.border = border

    wb.save(buffer)
    st.download_button(
        "\uc5d1\uc140 \ub2e4\uc6b4\ub85c\ub4dc",
        data=buffer.getvalue(),
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=key,
    )


def inject_theme() -> None:
    st.markdown(THEME_CSS, unsafe_allow_html=True)


def main() -> None:
    st.set_page_config(page_title="\uc218\uc8fc \ub300\uc2dc\ubcf4\ub4dc", layout="wide")
    inject_theme()
    st.title("\uc218\uc8fc \ub300\uc2dc\ubcf4\ub4dc")

    with st.sidebar:
        st.subheader("\ub370\uc774\ud130")
        upload = st.file_uploader("\uc5d1\uc140 \uc5c5\ub85c\ub4dc", type=["xlsx"])
        refresh = st.button("\ub370\uc774\ud130 \uc0c8\ub85c\uace0\uce68")
        if refresh:
            st.cache_data.clear()

    if upload:
        data = load_from_bytes(upload.getvalue(), upload.name)
        source_label = f"\uc5c5\ub85c\ub4dc \ud30c\uc77c: {upload.name}"
    else:
        if not DEFAULT_FILE.exists():
            st.error(f"\ub370\uc774\ud130 \ud30c\uc77c\uc744 \ucc3e\uc744 \uc218 \uc5c6\uc2b5\ub2c8\ub2e4: {DEFAULT_FILE}")
            st.stop()
        mtime = DEFAULT_FILE.stat().st_mtime
        data = load_from_path(str(DEFAULT_FILE), mtime)
        source_label = (
            f"\uae30\ubcf8 \ud30c\uc77c: {DEFAULT_FILE.name} "
            f"(\uc218\uc815: {datetime.fromtimestamp(mtime)})"
        )

    st.caption(source_label)

    tabs = st.tabs([TAB_ORDER_STATUS, TAB_BY_ITEM, TAB_PRODUCT_SUMMARY, TAB_ISSUES])
    shared_filters: dict | None = None

    with tabs[0]:
        st.subheader(TAB_ORDER_STATUS)
        df = data["order_status"].copy()
        month_range = render_period_controls(df, "main")

        detail_df, summary_df, shared_filters = apply_order_filters(
            df, month_range, show_sidebar=True
        )
        query = st.text_input(
            "\ud1b5\ud569 \uac80\uc0c9 (\uc6d0\ud558\ub294 \ud0a4\uc6cc\ub4dc\ub97c \uc785\ub825\ud558\uba74 \ud3ec\ud568\ub41c \ud589\ub9cc \ud45c\uc2dc\ub429\ub2c8\ub2e4)",
            "",
            key="main_search",
        )
        summary_df = apply_search(summary_df, query)
        render_year_summary(summary_df, "main")

        detail_df = apply_search(detail_df, query)
        detail_df = move_note_before_year(detail_df)

        numeric_cols = ORDER_STATUS_NUMERIC + [COL_YEAR]
        detail_df = detail_df.drop(columns=[SEARCH_COL], errors="ignore")
        display_df = prepare_display(
            detail_df,
            numeric_cols,
            ORDER_STATUS_DATE,
            ORDER_STATUS_MIXED_DATE,
            ORDER_STATUS_PERCENT,
        )
        use_status_style = len(display_df) <= MAX_STATUS_STYLE_ROWS
        styled = build_styler(
            display_df,
            numeric_cols,
            ORDER_STATUS_DATE + ORDER_STATUS_MIXED_DATE,
            COL_STATUS if use_status_style else None,
            ORDER_STATUS_PERCENT,
        )
        if not use_status_style:
            st.caption(
                "\ud589\uc774 \ub9ce\uc544 \ud604\uc7ac\uc0c1\ud0dc \uc0c9\uc0c1\ub9cc \uc0dd\ub7b5\ud588\uc2b5\ub2c8\ub2e4. \uac80\uc0c9 \ubc94\uc704\ub97c \uc904\uc774\uba74 \uc0c9\uc0c9\uc774 \uc801\uc6a9\ub429\ub2c8\ub2e4."
            )
        st.dataframe(styled, use_container_width=True, height=650)
        download_excel_button(
            detail_df,
            "order_status_filtered.xlsx",
            numeric_cols,
            ORDER_STATUS_DATE,
            key="main_download",
            mixed_date_cols=ORDER_STATUS_MIXED_DATE,
            percent_cols=ORDER_STATUS_PERCENT,
        )

    with tabs[1]:
        st.subheader(TAB_BY_ITEM)
        df = data["order_status_by_item"].copy()
        month_range = render_period_controls(df, "item")

        detail_df, summary_df, _ = apply_order_filters(
            df,
            month_range,
            filters=shared_filters or {},
            show_sidebar=False,
            apply_month_filter=False,
        )
        query = st.text_input(
            "\ud1b5\ud569 \uac80\uc0c9 (\uc6d0\ud558\ub294 \ud0a4\uc6cc\ub4dc\ub97c \uc785\ub825\ud558\uba74 \ud3ec\ud568\ub41c \ud589\ub9cc \ud45c\uc2dc\ub429\ub2c8\ub2e4)",
            "",
            key="item_search",
        )
        summary_df = apply_search(summary_df, query)
        render_year_summary(summary_df, "item")

        detail_df = apply_search(detail_df, query)
        detail_df = move_note_before_year(detail_df)

        numeric_cols = ORDER_STATUS_NUMERIC + [COL_YEAR]
        detail_df = detail_df.drop(columns=[SEARCH_COL], errors="ignore")
        display_df = prepare_display(
            detail_df,
            numeric_cols,
            ORDER_STATUS_DATE,
            ORDER_STATUS_MIXED_DATE,
            ORDER_STATUS_PERCENT,
        )
        use_status_style = len(display_df) <= MAX_STATUS_STYLE_ROWS
        styled = build_styler(
            display_df,
            numeric_cols,
            ORDER_STATUS_DATE + ORDER_STATUS_MIXED_DATE,
            COL_STATUS if use_status_style else None,
            ORDER_STATUS_PERCENT,
        )
        if not use_status_style:
            st.caption(
                "\ud589\uc774 \ub9ce\uc544 \ud604\uc7ac\uc0c1\ud0dc \uc0c9\uc0c1\ub9cc \uc0dd\ub7b5\ud588\uc2b5\ub2c8\ub2e4. \uac80\uc0c9 \ubc94\uc704\ub97c \uc904\uc774\uba74 \uc0c9\uc0c9\uc774 \uc801\uc6a9\ub429\ub2c8\ub2e4."
            )
        st.dataframe(styled, use_container_width=True, height=650)
        download_excel_button(
            detail_df,
            "order_status_by_item_filtered.xlsx",
            numeric_cols,
            ORDER_STATUS_DATE,
            key="item_download",
            mixed_date_cols=ORDER_STATUS_MIXED_DATE,
            percent_cols=ORDER_STATUS_PERCENT,
        )

    with tabs[2]:
        st.subheader(TAB_PRODUCT_SUMMARY)
        df = data["order_status_by_item"].copy()
        month_range = render_period_controls(df, "product")

        detail_df, _, _ = apply_order_filters(
            df,
            month_range,
            filters=shared_filters or {},
            show_sidebar=False,
            apply_month_filter=False,
        )
        query = st.text_input(
            "\ud1b5\ud569 \uac80\uc0c9 (\uc791\uc9c0\ubc88\ud638/\ud488\uba85 \ubc94\uc704\ub85c \uac80\uc0c9)",
            "",
            key="product_summary_search",
        )
        detail_df = apply_search(detail_df, query)

        if detail_df.empty:
            st.info("\ud574\ub2f9 \uae30\uac04\uc5d0 \ub370\uc774\ud130\uac00 \uc5c6\uc2b5\ub2c8\ub2e4.")
        else:
            summary_df = compute_product_priority(detail_df)
            numeric_cols = [
                COL_PRIORITY,
                COL_AVG_DEMAND,
                COL_TOTAL_QTY,
                COL_PO_COUNT,
                COL_PO_STREAK,
            ]
            styled = build_styler(
                summary_df,
                numeric_cols,
                [],
                status_col=None,
                percent_cols=[COL_SHARE],
            )
            styled = apply_styler_widths(styled, list(summary_df.columns))
            st.dataframe(styled, use_container_width=True, height=560)

    with tabs[3]:
        st.subheader(TAB_ISSUES)
        issues = data["order_status_by_item"].copy()
        if COL_NOTE not in issues.columns:
            st.info("\ud2b9\uc774\uc0ac\ud56d \ub370\uc774\ud130\uac00 \uc5c6\uc2b5\ub2c8\ub2e4.")
            return
        issues = issues[issues[COL_NOTE].notna()]
        issues = issues[issues[COL_NOTE].astype(str).str.strip().ne("")]
        if issues.empty:
            st.info("\ud2b9\uc774\uc0ac\ud56d \ub370\uc774\ud130\uac00 \uc5c6\uc2b5\ub2c8\ub2e4.")
            return

        base_cols = [COL_MONTH, COL_TYPE, COL_WORKNO, COL_CUSTOMER, COL_PRODUCT, COL_NOTE]
        issues = issues[base_cols].drop_duplicates().copy()
        issues[COL_ISSUE_KEY] = build_issue_key(issues)

        tracker = load_issue_tracker(ISSUE_TRACKER_PATH)
        merged = issues.merge(tracker, on=COL_ISSUE_KEY, how="left")
        merged[COL_RESOLVED] = merged[COL_RESOLVED].fillna(False).astype(bool)
        merged[COL_CLOSED_DATE] = pd.to_datetime(
            merged[COL_CLOSED_DATE], errors="coerce"
        ).dt.date
        merged[COL_ISSUE_DATE] = pd.to_datetime(
            merged[COL_ISSUE_DATE], errors="coerce"
        ).dt.date

        merged = add_search_column(merged)
        query = st.text_input(
            "\ud1b5\ud569 \uac80\uc0c9 (\ud2b9\uc774\uc0ac\ud56d \ubaa8\ub4e0 \ud56d\ubaa9\uc5d0\uc11c \uac80\uc0c9)",
            "",
            key="issue_search",
        )
        merged = apply_search(merged, query).drop(columns=[SEARCH_COL], errors="ignore")

        unresolved = merged[~merged[COL_RESOLVED]].copy()
        resolved = merged[merged[COL_RESOLVED]].copy()

        st.caption(
            f"\ucd1d {len(merged):,}\uac74 \u00b7 \ubbf8\ud574\uacb0 {len(unresolved):,}\uac74 \u00b7 \uc885\uacb0 {len(resolved):,}\uac74"
        )

        display_cols = base_cols + [COL_ISSUE_DATE, COL_RESOLVED, COL_CLOSED_DATE]
        st.markdown("**\ubbf8\ud574\uacb0 \uc548\uac74**")
        editor_df = unresolved[display_cols].copy()
        if st.button("\uc804\uccb4 \ud574\uacb0", key="issue_resolve_all"):
            editor_df[COL_RESOLVED] = True
            editor_df[COL_CLOSED_DATE] = date.today()
        edited = st.data_editor(
            editor_df,
            use_container_width=True,
            height=calc_table_height(
                len(editor_df),
                row_height=ISSUE_ROW_HEIGHT,
                max_height=ISSUE_TABLE_MAX_HEIGHT,
            ),
            row_height=ISSUE_ROW_HEIGHT,
            num_rows="fixed",
            column_config={
                COL_RESOLVED: st.column_config.CheckboxColumn(
                    "\ud574\uacb0\uc5ec\ubd80"
                ),
                COL_CLOSED_DATE: st.column_config.DateColumn(
                    "\uc885\uacb0\uc77c", format="iso8601"
                ),
                COL_ISSUE_DATE: st.column_config.DateColumn(
                    "\uc548\uac74\uc0c1\uc815\uc77c", format="iso8601"
                ),
            },
            key="issue_editor",
        )
        if st.button("\uc800\uc7a5", key="issue_save"):
            updated = pd.concat([edited, resolved[display_cols]], ignore_index=True)
            updated[COL_RESOLVED] = updated[COL_RESOLVED].fillna(False).astype(bool)
            updated[COL_CLOSED_DATE] = pd.to_datetime(
                updated[COL_CLOSED_DATE], errors="coerce"
            ).dt.date
            updated[COL_ISSUE_DATE] = pd.to_datetime(
                updated[COL_ISSUE_DATE], errors="coerce"
            ).dt.date
            today = date.today()
            missing_closed = updated[COL_CLOSED_DATE].isna() | (
                updated[COL_CLOSED_DATE].astype(str).str.strip() == ""
            )
            updated.loc[
                updated[COL_RESOLVED] & missing_closed,
                COL_CLOSED_DATE,
            ] = today
            updated.loc[~updated[COL_RESOLVED], COL_CLOSED_DATE] = pd.NaT
            updated[COL_ISSUE_KEY] = build_issue_key(updated)
            tracker_out = (
                updated[[COL_ISSUE_KEY, COL_RESOLVED, COL_CLOSED_DATE, COL_ISSUE_DATE]]
                .drop_duplicates(subset=[COL_ISSUE_KEY], keep="last")
                .reset_index(drop=True)
            )
            save_issue_tracker(tracker_out, ISSUE_TRACKER_PATH)
            st.success("\uc800\uc7a5\ud588\uc2b5\ub2c8\ub2e4.")

        st.markdown("**\uc885\uacb0 \uc548\uac74**")
        if resolved.empty:
            st.caption("\uc885\uacb0\ub41c \uc548\uac74\uc774 \uc5c6\uc2b5\ub2c8\ub2e4.")
        else:
            st.data_editor(
                resolved[display_cols],
                use_container_width=True,
                height=calc_table_height(
                    len(resolved),
                    row_height=ISSUE_ROW_HEIGHT,
                    max_height=ISSUE_RESOLVED_MAX_HEIGHT,
                ),
                row_height=ISSUE_ROW_HEIGHT,
                disabled=True,
            )


if __name__ == "__main__":
    main()
